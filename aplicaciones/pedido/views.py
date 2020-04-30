from django.shortcuts import render, redirect
from django.views.generic import CreateView, ListView, UpdateView, DeleteView, TemplateView, DetailView, View
from aplicaciones.pedido.models import *
from aplicaciones.pedido.forms import *  
from django.urls import reverse_lazy, reverse
from aplicaciones.empresa.eliminaciones import get_deleted_objects
from django.contrib.auth.decorators import permission_required
from django.utils.decorators import method_decorator
from django.contrib.auth.mixins import LoginRequiredMixin

from openpyxl.styles import Font, Fill, Alignment
from django.http import HttpResponse
from openpyxl import Workbook
from django.db.models import Q
from django.http import JsonResponse
from django.contrib import messages

from django.conf import settings
from io import BytesIO
# //////////////////////////////////////
from reportlab.pdfgen import canvas
from reportlab.platypus import Paragraph, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
# ////////////////////////////////////////
import os
#Librerias reportlab a usar:
from reportlab.platypus import (SimpleDocTemplate, PageBreak, Image, Spacer,
Paragraph, Table, TableStyle)
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4, letter, landscape
from reportlab.lib import colors


class ProductoLista(LoginRequiredMixin, ListView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = Producto
    paginate_by = 250
    template_name = 'pedido/productos.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user
        return context

    def get_queryset(self):
        queryset = super(ProductoLista, self).get_queryset()
        buesqueda = self.request.GET.get('search')
        if buesqueda != None:
            queryset = queryset.filter(Q(prod_codigo=buesqueda) | Q(
                prod_descripcion__icontains=buesqueda))

        return queryset


class PrductoCreate(LoginRequiredMixin, CreateView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = Producto
    form_class = ProductoForm
    template_name = 'pedido/crear_producto.html'
    success_url = reverse_lazy('pedido:ProductoLista')

    @method_decorator(permission_required('pedido.add_producto', reverse_lazy('requiere_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(PrductoCreate, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user
        return context


class ProductoUpdate(LoginRequiredMixin, UpdateView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = Producto
    form_class = ProductoForm
    template_name = 'pedido/crear_producto.html'
    success_url = reverse_lazy('pedido:ProductoLista')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user
        return context

    @method_decorator(permission_required('pedido.change_producto', reverse_lazy('requiere_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(ProductoUpdate, self).dispatch(*args, **kwargs)


class ProductoDelete(LoginRequiredMixin, DeleteView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = Producto
    template_name = 'eliminaciones.html'
    success_url = reverse_lazy('pedido:ProductoLista')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user
        deletable_objects, model_count, protected = get_deleted_objects([
                                                                        self.object])
        context['deletable_objects'] = deletable_objects
        context['model_count'] = dict(model_count).items()
        context['protected'] = protected
        return context

    @method_decorator(permission_required('pedido.delete_producto', reverse_lazy('requiere_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(ProductoDelete, self).dispatch(*args, **kwargs)


class PedidoList(LoginRequiredMixin, ListView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = Pedido
    template_name = 'pedido/pedido_list.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user
        return context

    def get_queryset(self): 

        queryset = super(PedidoList, self).get_queryset()
        status = self.request.GET.get('status')
        tipo_pedido = self.request.GET.get('tipo_pedido')
        inicio = self.request.GET.get('inicio')
        fin = self.request.GET.get('fin')        

        if self.request.user.is_superuser or self.request.user.tipo_user == 2:
            if len(self.request.GET) == 0:
                queryset = queryset.filter(ped_estatusPedido=1)

            if status != '0':
                queryset = queryset.filter(ped_estatusPedido=status)
            
            if inicio !=None and fin != None:
                queryset=queryset.filter(ped_fechaCreacion__range=(inicio, fin))
            
            if tipo_pedido != '0':
                queryset = queryset.filter(dtl_tipo_pedido=tipo_pedido)
                    
                
         
        elif self.request.user.tipo_user == 1:
            id_zona = self.request.user.zona_pertene
            queryset = queryset.filter(ped_id_Suc__suc_zona=id_zona)

            if len(self.request.GET) == 0:
                queryset = queryset.filter(ped_estatusPedido=1)
            if inicio !=None and fin != None:
                queryset=queryset.filter(ped_fechaCreacion__range=(inicio, fin))
            if status != '0':
                queryset = queryset.filter(ped_estatusPedido=status)
                    
            if tipo_pedido != '0':
                queryset = queryset.filter(dtl_tipo_pedido=tipo_pedido)
                    

        elif self.request.user.tipo_user == 3:
            id_suc = self.request.user.suc_pertene
            queryset = queryset.filter(ped_id_Suc=id_suc)

            if len(self.request.GET) == 0:
                queryset = queryset.filter(ped_estatusPedido=1)
            if inicio !=None and fin != None:
                queryset=queryset.filter(ped_fechaCreacion__range=(inicio, fin))
            if status != '0':
                queryset = queryset.filter(ped_estatusPedido=status)
            if tipo_pedido != '0':
                queryset = queryset.filter(dtl_tipo_pedido=tipo_pedido)
                    

        
        return queryset



class ReporteDetallePedido(View):
    def get(self, request, *args, **kwargs):
        from django.contrib.humanize.templatetags.humanize import intcomma
        print ("Genero el PDF")
        response = HttpResponse(content_type='application/pdf')
        pdf_name = "clientes.pdf"  # llamado clientes
        # la linea 26 es por si deseas descargar el pdf a tu computadora
        # response['Content-Disposition'] = 'attachment; filename=%s' % pdf_name
        buff = BytesIO()
        doc = SimpleDocTemplate(buff,
                                pagesize=letter,
                                rightMargin=40,
                                leftMargin=40,
                                topMargin=60,
                                bottomMargin=18,
                                )
        items = []
        styles = getSampleStyleSheet()
        total_detalles=DetallePedido.objects.filter(dtl_id_pedido=request.GET.get('pedido_id')).aggregate(total=Sum(F('dtl_cantidad')* F('dtl_precio'), output_field=FloatField()))
        
        header = Paragraph("DETALLE DE PEDIDO N°"+request.GET.get('pedido_id'), styles['Heading1'])
        pedido_get=Pedido.objects.get(ped_id_ped=request.GET.get('pedido_id'))
        sucursal_num = pedido_get.ped_id_Suc.suc_numero
        sucursal_direccion = pedido_get.ped_id_Suc.suc_direccion

        txt_sucursal_num=Paragraph('N° de Sucursal: '+'N/A' if sucursal_num == None else  'N° de Sucursal: '+sucursal_num, styles['Heading4'])
        txt_sucursal_direccion=Paragraph('Direccion: '+'N/A' if sucursal_direccion == None else 'Direccion: '+sucursal_direccion, styles['Heading4'])
        items.append(header)
        items.append(txt_sucursal_num)
        items.append(txt_sucursal_direccion)
        headings = ('Codigo', 'Descripción', 'Cantidad', 'Precio', 'Subtotal')
        query_result = [(p.dtl_codigo, Paragraph(p.dtl_descripcion, styles['BodyText']), p.dtl_cantidad,p.dtl_precio, round(p.dtl_cantidad*p.dtl_precio, 2)) for p in DetallePedido.objects.filter(dtl_id_pedido=request.GET.get('pedido_id'))]
        
        total=0
        if total_detalles['total'] != None:
            total=round(total_detalles['total'], 2)
    
            

        final_line=[('','','',Paragraph("Total", styles['Heading5']),Paragraph(intcomma(total), styles['Heading5']))]

        t = Table([headings] + query_result+final_line, colWidths=[2.5 * cm, 10 * cm, 2 * cm, 2 * cm, 2 * cm])
        t.setStyle(TableStyle(
            [
                ('GRID', (0, 0), (4, -1), 1, colors.dodgerblue),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
                ('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue)
            ]
        ))
        items.append(t)
        doc.build(items)
        response.write(buff.getvalue())
        buff.close()
        return response
        


class CarritoLista(LoginRequiredMixin, ListView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = DetallePedido
    template_name = 'pedido/carrito.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user
        context['ped_papeleria'] = DetallePedido.objects.filter(
            dtl_creado_por=self.request.user, dtl_status=False, dtl_tipo_pedido=1)
        context['ped_lim_consultorio'] = DetallePedido.objects.filter(
            dtl_creado_por=self.request.user, dtl_status=False, dtl_tipo_pedido=3)
        context['ped_lim_sucursal'] = DetallePedido.objects.filter(
            dtl_creado_por=self.request.user, dtl_status=False, dtl_tipo_pedido=2)
        context['ped_consumibles'] = DetallePedido.objects.filter(
            dtl_creado_por=self.request.user, dtl_status=False, dtl_tipo_pedido=4)

        return context

    def get_queryset(self):
        queryset = super(CarritoLista, self).get_queryset()
        queryset = queryset.filter(
            dtl_creado_por=self.request.user, dtl_status=False)
        return queryset

    def post(self, request, *args, **kwargs):
        from datetime import datetime, date 
        import calendar
        # ESTABLECEMOS LA FECHA ACTUAL
        today = datetime.now()
        # CONSULTAMOS CUAL ES EL ULTIMO DIA DEL MES ACTUAL
        last_day=calendar.monthrange(today.year, today.month)[1]
        # INICIALIZAMOS LA FECHA INICIAL
        start_date = datetime(today.year, today.month, 1)
        # INICIALIZAMOS LA FECHA FINAL
        end_date = datetime(today.year, today.month, last_day)

        tipo_pedido = request.POST.get('tipo_pedido')
        # pedido_count=Pedido.objects.filter(dtl_tipo_pedido=1, ped_id_Suc=self.request.user.suc_pertene, ped_fechaCreacion__range=(start_date,end_date)).count()

        if tipo_pedido == 'papeleria':
            pedido = Pedido(
                ped_id_Suc=self.request.user.suc_pertene,
                ped_id_UsuarioCreo=self.request.user,
                dtl_tipo_pedido=1,
                pedido_tipo_insumo=800044563,
            )
            pedido_count=Pedido.objects.filter(dtl_tipo_pedido=1, ped_id_Suc=self.request.user.suc_pertene, ped_fechaCreacion__range=(start_date,end_date)).count()
            if pedido_count == 0:
                pedido.save()
                DetallePedido.objects.filter(dtl_creado_por=self.request.user,dtl_status=False, dtl_tipo_pedido=1).update(dtl_status=True, dtl_id_pedido=pedido)
            
        if tipo_pedido == 'sucursal':
            pedido = Pedido(
                ped_id_Suc=self.request.user.suc_pertene,
                ped_id_UsuarioCreo=self.request.user,
                dtl_tipo_pedido=2,
            )
            pedido_count=Pedido.objects.filter(dtl_tipo_pedido=2, ped_id_Suc=self.request.user.suc_pertene, ped_fechaCreacion__range=(start_date,end_date)).count()
            if pedido_count == 0:
                pedido.save()
                DetallePedido.objects.filter(dtl_creado_por=self.request.user, dtl_status=False, dtl_tipo_pedido=2).update(dtl_status=True, dtl_id_pedido=pedido)
            
        if tipo_pedido == 'consultorio':
            pedido = Pedido(
                ped_id_Suc=self.request.user.suc_pertene,
                ped_id_UsuarioCreo=self.request.user,
                dtl_tipo_pedido=3,
            )
            pedido_count=Pedido.objects.filter(dtl_tipo_pedido=3, ped_id_Suc=self.request.user.suc_pertene, ped_fechaCreacion__range=(start_date,end_date)).count()
            if pedido_count == 0:
                pedido.save()
                DetallePedido.objects.filter(dtl_creado_por=self.request.user, dtl_status=False, dtl_tipo_pedido=3).update(dtl_status=True, dtl_id_pedido=pedido)

        if tipo_pedido == 'consumible':
            pedido = Pedido(
                ped_id_Suc=self.request.user.suc_pertene,
                ped_id_UsuarioCreo=self.request.user,
                dtl_tipo_pedido=4,
            )
            pedido_count=Pedido.objects.filter(dtl_tipo_pedido=4, ped_id_Suc=self.request.user.suc_pertene, ped_fechaCreacion__range=(start_date,end_date)).count()
            if pedido_count == 0:
                pedido.save()
                DetallePedido.objects.filter(dtl_creado_por=self.request.user, dtl_status=False, dtl_tipo_pedido=4).update(dtl_status=True, dtl_id_pedido=pedido)

        url = reverse_lazy('pedido:carrito')
        return redirect(url)


class CarritoDelete(LoginRequiredMixin, DeleteView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = DetallePedido
    template_name = 'eliminaciones.html'
    success_url = reverse_lazy('pedido:carrito')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user
        deletable_objects, model_count, protected = get_deleted_objects([
                                                                        self.object])
        context['deletable_objects'] = deletable_objects
        context['model_count'] = dict(model_count).items()
        context['protected'] = protected
        return context


class DowloadExcelPedido(LoginRequiredMixin, TemplateView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'

    def get(self, request, *args, **kwargs):
        from openpyxl.utils import get_column_letter
        import datetime
        wb = Workbook()
        ws = wb.active
        id_pedido = self.kwargs.get('pk')
        ped = Pedido.objects.get(ped_id_ped=id_pedido)

        ws['A1'] = 'Detalle Pedido '+str(ped.ped_id_Suc)+' N°' + str(id_pedido)
        st = ws['A1']
        st.font = Font(size=14, b=True, color="004ee0")
        st.alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:F1')
        ws.sheet_properties.tabColor = "1072BA"

        ws['A2'] = 'Producto'
        ws['B2'] = 'Descripcion'
        ws['C2'] = 'Sucursal'
        ws['D2'] = 'Cantidad'
        ws['E2'] = 'Precio'
        ws['F2'] = 'Total'
        cont = 3
        detalle = DetallePedido.objects.filter(dtl_id_pedido=id_pedido)
        for cto in detalle:
            ws.cell(row=cont, column=1).value = str(cto.dtl_codigo)
            ws.cell(row=cont, column=2).value = str(cto.dtl_descripcion)
            ws.cell(row=cont, column=3).value = str(
                cto.dtl_id_pedido.ped_id_Suc)
            ws.cell(row=cont, column=4).value = cto.dtl_cantidad
            ws.cell(row=cont, column=5).value = cto.dtl_precio
            ws.cell(row=cont, column=6).value = (
                cto.dtl_cantidad * cto.dtl_precio)
            ws.cell(row=cont, column=6).number_format = '#,##0'

            cont += 1

        ws["F"+str(cont)] = "=SUM(F3:F"+str(cont-1)+")"
        ws["F"+str(cont)].number_format = '#,##0'

        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    if cell.row != 1:
                        dims[cell.column] = max(
                            (dims.get(cell.column, 0), len(str(cell.value))))

        for col, value in dims.items():
            ws.column_dimensions[get_column_letter(col)].width = value

        ahora = datetime.datetime.now()
        Pedido.objects.filter(ped_id_ped=id_pedido).update(ped_estatusPedido=6, ped_id_usr_descargo_excel=self.request.user)
        nombre_archivo = 'pedido_no_'+str(id_pedido)+'.xls'
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class DetallePedidolit(LoginRequiredMixin, TemplateView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    template_name = 'pedido/ver_detalle.html'  

    def get_context_data(self, **kwargs):
        from django.db.models import Sum
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user
        id_pedido = self.kwargs.get('pk')
        context['articulos'] = DetallePedido.objects.filter(dtl_id_pedido=id_pedido)
        context['suma_total'] = DetallePedido.objects.filter(dtl_id_pedido=id_pedido).aggregate(total= Sum(F('dtl_precio')*F('dtl_cantidad'), output_field=FloatField()))['total']
        context['suma_total_partids'] = DetallePedido.objects.filter(dtl_id_pedido=id_pedido).aggregate(Sum('dtl_cantidad'))
        print(context['suma_total'])
        return context 


class DetallePeditoEliminar(DeleteView):
    model = DetallePedido
    template_name = 'eliminaciones.html'
    success_url = reverse_lazy('pedido:detalle_pedido')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user
        return context

    def get_success_url(self):
        slug = self.kwargs['pedo_id']
        link = reverse_lazy('pedido:detalle_pedido', kwargs={'pk': slug})
        return link

    @method_decorator(permission_required('pedido.delete_detallepedido', reverse_lazy('requiere_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(DetallePeditoEliminar, self).dispatch(*args, **kwargs)


class DetallePedidoEdit(UpdateView):
    model = DetallePedido
    template_name = 'pedido/conf_create.html'
    form_class = DetallePedidoFormEdit 
    success_url = '/'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user
        return context

    def form_valid(self, form):
        pedido_id=self.kwargs['pedo_id'] 
        obj_pedido=Pedido.objects.get(ped_id_ped=pedido_id)  
        obj_detalle_p=DetallePedido.objects.get(dtl_id_detalle=self.kwargs['pk'])     
        tipo_ped=obj_detalle_p.dtl_tipo_pedido

        producto_obj = Producto.objects.get(prod_codigo=obj_detalle_p.dtl_codigo)

        limite_credito=0
        tipo_pedido=''
        if tipo_ped == 1:
            # Papeleria
            tipo_pedido='Papeleria'
            limite_credito = obj_pedido.ped_id_UsuarioCreo.suc_pertene.suc_monto_papeleria
        elif tipo_ped == 2:
            # Limpieza
            tipo_pedido='Limpieza'
            limite_credito = obj_pedido.ped_id_UsuarioCreo.suc_pertene.suc_monto_limpieza
        elif tipo_ped == 3:
            # Limpieza Consultorio
            tipo_pedido='Limpieza Consultorio'
            limite_credito = obj_pedido.ped_id_UsuarioCreo.suc_pertene.suc_monto_limpieza_oficina
        elif tipo_ped == 4:
            # Toner
            tipo_pedido='Toner'
            limite_credito = obj_pedido.ped_id_UsuarioCreo.suc_pertene.suc_monto_consumible
        
        total_ped=DetallePedido.objects.filter(dtl_id_pedido=self.kwargs['pedo_id']).exclude(dtl_id_detalle=self.kwargs['pk']).aggregate(total_venta=Sum(F('dtl_cantidad') * F('dtl_precio'), output_field=FloatField()))['total_venta']
        total_ped = 0 if total_ped == None else total_ped

        cantidad=form.instance.dtl_cantidad
        subtotal_producto = cantidad * producto_obj.prod_precio

        total = total_ped + subtotal_producto

        if total > limite_credito:
            messages.warning(self.request, 
            'Pedido:{} con monto maximo:{} supera el monto con:{} reconsidere actualizar con menos producto o dejar tal cual está.'
            .format(obj_pedido, limite_credito, total))
            url = reverse_lazy('pedido:detalle_pedido_update', kwargs={'pk':self.kwargs['pk'], 'pedo_id':self.kwargs['pedo_id']})
            return redirect(url)

        return super().form_valid(form)

    def get_success_url(self):
        slug = self.kwargs['pedo_id']
        link = reverse_lazy('pedido:detalle_pedido', kwargs={'pk': slug})
        return link

    @method_decorator(permission_required('pedido.change_detallepedido', reverse_lazy('requiere_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(DetallePedidoEdit, self).dispatch(*args, **kwargs)


class AutorizarPedido(LoginRequiredMixin, TemplateView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    template_name = 'pedido/ver_detalle.html'

    def get(self, request, *args, **kwargs):
        from datetime import datetime
        from django.contrib.humanize.templatetags.humanize import naturalday
        ahora = datetime.now()
        id_pedido = self.kwargs.get('pk')
        confs = Configuracion_pedido.objects.all()
        for config in confs:
            if ahora.date() <= config.conf_fecha_fin_autorizador:
                Pedido.objects.filter(ped_id_ped=id_pedido).update(
                    ped_estatusPedido=2, ped_id_UsuarioAutorizo=self.request.user, ped_fechaAutorizacion=ahora)
            else:
                messages.error(request, 'No puede autorizar fuera de tiempo, fecha limite '+ naturalday(config.conf_fecha_fin_autorizador))


        rev = reverse_lazy('pedido:listar_pedido')
        get_encode=self.request.GET.urlencode()
        rev = rev +'?'+get_encode
        return redirect(rev) 

    @method_decorator(permission_required('pedido.change_pedido', reverse_lazy('requiere_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(AutorizarPedido, self).dispatch(*args, **kwargs)


class RechazarPedido(LoginRequiredMixin, TemplateView): 
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    template_name = 'pedido/ver_detalle.html'

    def get(self, request, *args, **kwargs):
        import datetime
        ahora = datetime.datetime.now()
        id_pedido = self.kwargs.get('pk')
        Pedido.objects.filter(ped_id_ped=id_pedido).update(
            ped_estatusPedido=4, ped_id_UsuarioCancelo=self.request.user, ped_fechaCancelacion=ahora)

        rev = reverse_lazy('pedido:listar_pedido')
        return redirect(rev+'?'+self.request.GET.urlencode())

    @method_decorator(permission_required('pedido.change_pedido', reverse_lazy('requiere_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(RechazarPedido, self).dispatch(*args, **kwargs)


class PedidoUpdate(LoginRequiredMixin, UpdateView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = Pedido
    form_class = PedidoForm
    template_name = 'pedido/pedido_form.html'
    success_url = reverse_lazy('pedido:listar_pedido')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user
        return context

    def form_valid(self, form):
        import datetime
        ahora = datetime.datetime.now()
        """If the form is valid, save the associated model."""
        self.object = form.save()
        self.object.ped_estatusPedido = 3
        self.object.ped_fechaSubidaFac = ahora
        return super().form_valid(form)

    @method_decorator(permission_required('pedido.change_pedido', reverse_lazy('requiere_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(PedidoUpdate, self).dispatch(*args, **kwargs)


class PedidoDetalleList(LoginRequiredMixin, DetailView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = Pedido
    template_name = 'pedido/detalle_ped.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user
        return context


class ConfiguracionList(ListView):
    model = Configuracion_pedido
    template_name = 'pedido/conf_pedido.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user
        return context


class ConfigCreate(CreateView):
    model = Configuracion_pedido
    form_class = ConfForm
    template_name = 'pedido/conf_create.html'
    success_url = reverse_lazy('pedido:configuracion_pedido')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user
        return context


class ConfigUpdate(UpdateView):
    model = Configuracion_pedido
    form_class = ConfFormEdit
    template_name = 'pedido/conf_create.html'
    success_url = reverse_lazy('pedido:configuracion_pedido')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user
        return context


class PedidoDeleteRechazado(TemplateView):
    template_name = 'eliminacion_masiva.html'

    def post(self, request, *args, **kwargs):
        Pedido.objects.filter(ped_estatusPedido=4).delete()
        url = reverse_lazy('pedido:listar_pedido')
        return redirect(url)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuario'] = self.request.user
        context['dell_masivo_list'] = Pedido.objects.filter(
            ped_estatusPedido=4)

        return context

    @method_decorator(permission_required('pedido.delete_pedido', reverse_lazy('requiere_permisos')))
    def dispatch(self, *args, **kwargs):
        return super(PedidoDeleteRechazado, self).dispatch(*args, **kwargs)


class GeneraValuesJsonPedidos(View):
    def post(self, request, *args, **kwargs):
        import json
        from datetime import datetime
        body = json.loads(request.body)
        inicio = body.get('fecha_inicio')
        fin =body.get('fecha_fin')

        if inicio == '' or fin == '':
            data = {
                'msn': 'Debe de indicar las dos fechas',
                'status': False,
            }
            return JsonResponse(data, status=400)
        else:
            inicio = datetime.strptime(body.get('fecha_inicio'), '%Y-%m-%d')
            fin = datetime.strptime(body.get('fecha_fin'), '%Y-%m-%d')
            if fin > inicio:
                query=Pedido.objects.filter(ped_fechaCreacion__range=[inicio, fin]).values(
                    'ped_id_ped',
                    'ped_id_Suc__suc_numero',
                    'ped_id_Suc__suc_nombre',
                    'ped_id_Suc__suc_direccion',
                    'pedido_tipo_insumo',
                )
                lista=list(query)
                data = {
                    'obj_list': lista,
                    'status': True,
                }
                return JsonResponse(data)
            else:
                data = {
                    'msn': 'La fecha inicial no puede ser mayor o igual a la fecha final',
                    'status': False,
                }
                return JsonResponse(data, status=400)


class ReportPedido(TemplateView): 
    template_name = 'pedido/report_gen.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        inicio=self.request.GET.get('inicio')
        fin=self.request.GET.get('fin')
        tipo_pedido=self.request.GET.get('tipo_pedido')
        status=self.request.GET.getlist('status')

        query = DetallePedido.objects.filter(dtl_id_pedido__ped_fechaCreacion__range=(inicio, fin), dtl_id_pedido__ped_estatusPedido__in=status)
        if tipo_pedido != '0':
            query = query.filter(dtl_id_pedido__dtl_tipo_pedido=tipo_pedido)
        
        context['obj_list']=query
                
            
        return context
class DowloadReport(View):
    def get(self, request, *args, **kwargs):
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import PatternFill, fills
        import datetime
        from django.contrib.humanize.templatetags.humanize import naturalday
        TIPO_PEDIDO={0:"Todo", 1:"Papeleria", 2:"Limpieza", 3:"Limpieza Consultorio", 4:'Consumibles'}
        wb = Workbook()
        ws = wb.active
        
        tipo_pedido=self.request.GET.get('tipo_pedido')
        status=self.request.GET.getlist('status')
        
        inicio = datetime.datetime.strptime(request.GET.get('inicio'), '%Y-%m-%d').date()
        fin = datetime.datetime.strptime(request.GET.get('fin'), '%Y-%m-%d').date()

        ped_list = DetallePedido.objects.filter(dtl_id_pedido__ped_estatusPedido__in=status).values(
        'dtl_id_pedido', 
        'dtl_id_pedido__ped_fechaCreacion',
        'dtl_id_pedido__ped_id_Suc__suc_nombre',
        'dtl_id_pedido__ped_id_Suc__suc_zona__zona_nombre',
        'dtl_id_pedido__ped_id_UsuarioCreo__username',
        'dtl_id_pedido__dtl_tipo_pedido',
        'dtl_id_pedido__ped_id_UsuarioAutorizo__username'
        ).annotate(total_vent=Sum(F('dtl_cantidad') * F('dtl_precio'), output_field=FloatField()))
        if tipo_pedido == '0':
            ped_list=ped_list.filter(dtl_id_pedido__ped_fechaCreacion__range=(inicio, fin))
        else:
            ped_list=ped_list.filter(dtl_id_pedido__ped_fechaCreacion__range=(inicio, fin), dtl_id_pedido__dtl_tipo_pedido=tipo_pedido)
        

        ws['A1'] = 'Reporte de Pedidos '+naturalday(inicio)+' al '+naturalday(fin)
        st = ws['A1']
        st.font = Font(size=14, b=True, color="004ee0")
        

        st.alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:H1')
        ws.sheet_properties.tabColor = "1072BA"

        ws['A2'] = 'No. Pedido'
        ws['B2'] = 'Fecha Pedido'
        ws['C2'] = 'Sucursal'
        ws['D2'] = 'Zona'
        ws['E2'] = 'Usuario'
        ws['F2'] = 'Tipo'
        ws['G2'] = 'Autorizó'
        ws['H2'] = 'Total'

        ws['A2'].fill = PatternFill(start_color='007EDD', end_color='007EDD', fill_type = fills.FILL_PATTERN_LIGHTHORIZONTAL)
        ws['B2'].fill = PatternFill(start_color='007EDD', end_color='007EDD', fill_type = fills.FILL_PATTERN_LIGHTHORIZONTAL)
        ws['C2'].fill = PatternFill(start_color='007EDD', end_color='007EDD', fill_type = fills.FILL_PATTERN_LIGHTHORIZONTAL)
        ws['D2'].fill = PatternFill(start_color='007EDD', end_color='007EDD', fill_type = fills.FILL_PATTERN_LIGHTHORIZONTAL)
        ws['E2'].fill = PatternFill(start_color='007EDD', end_color='007EDD', fill_type = fills.FILL_PATTERN_LIGHTHORIZONTAL)
        ws['F2'].fill = PatternFill(start_color='007EDD', end_color='007EDD', fill_type = fills.FILL_PATTERN_LIGHTHORIZONTAL)
        ws['G2'].fill = PatternFill(start_color='007EDD', end_color='007EDD', fill_type = fills.FILL_PATTERN_LIGHTHORIZONTAL)
        ws['H2'].fill = PatternFill(start_color='007EDD', end_color='007EDD', fill_type = fills.FILL_PATTERN_LIGHTHORIZONTAL)


        
        cont = 3
        sub_total_gobal = 0
        for pedido in ped_list:
            ws.cell(row=cont, column=1).value = pedido['dtl_id_pedido']
            ws.cell(row=cont, column=2).value = str(naturalday(pedido['dtl_id_pedido__ped_fechaCreacion']))
            ws.cell(row=cont, column=3).value = str(pedido['dtl_id_pedido__ped_id_Suc__suc_nombre'])
            ws.cell(row=cont, column=4).value = str(pedido['dtl_id_pedido__ped_id_Suc__suc_zona__zona_nombre'])
            ws.cell(row=cont, column=5).value = str(pedido['dtl_id_pedido__ped_id_UsuarioCreo__username'])
            ws.cell(row=cont, column=6).value = str(TIPO_PEDIDO[pedido['dtl_id_pedido__dtl_tipo_pedido']])
            ws.cell(row=cont, column=7).value = str(pedido['dtl_id_pedido__ped_id_UsuarioAutorizo__username'])
            ws.cell(row=cont, column=8).value = pedido['total_vent']
            ws.cell(row=cont, column=8).number_format = '#,##0'

            sub_total_gobal += pedido['total_vent']

            ws.cell(row=cont, column=1).fill = PatternFill(start_color='007EDD', end_color='39A5F6', fill_type = fills.FILL_PATTERN_LIGHTHORIZONTAL)
            ws.cell(row=cont, column=2).fill = PatternFill(start_color='007EDD', end_color='39A5F6', fill_type = fills.FILL_PATTERN_LIGHTHORIZONTAL)
            ws.cell(row=cont, column=3).fill = PatternFill(start_color='007EDD', end_color='39A5F6', fill_type = fills.FILL_PATTERN_LIGHTHORIZONTAL)
            ws.cell(row=cont, column=4).fill = PatternFill(start_color='007EDD', end_color='39A5F6', fill_type = fills.FILL_PATTERN_LIGHTHORIZONTAL)
            ws.cell(row=cont, column=5).fill = PatternFill(start_color='007EDD', end_color='39A5F6', fill_type = fills.FILL_PATTERN_LIGHTHORIZONTAL)
            ws.cell(row=cont, column=6).fill = PatternFill(start_color='007EDD', end_color='39A5F6', fill_type = fills.FILL_PATTERN_LIGHTHORIZONTAL)
            ws.cell(row=cont, column=7).fill = PatternFill(start_color='007EDD', end_color='39A5F6', fill_type = fills.FILL_PATTERN_LIGHTHORIZONTAL)
            ws.cell(row=cont, column=8).fill = PatternFill(start_color='007EDD', end_color='39A5F6', fill_type = fills.FILL_PATTERN_LIGHTHORIZONTAL)
            
            cont += 1
            ws['A'+str(cont)] = '##'
            ws['B'+str(cont)] = 'Sucursal'
            ws['C'+str(cont)] = 'Zona'
            ws['D'+str(cont)] = 'Producto'
            ws['E'+str(cont)] = 'Descripcion'
            ws['F'+str(cont)] = 'Cantidad'
            ws['G'+str(cont)] = 'Precio'
            ws['H'+str(cont)] = 'SubTotal'

            ws['A'+str(cont)].fill = PatternFill(start_color='FFEE08', end_color='FFEE08', fill_type = 'solid')
            ws['B'+str(cont)].fill = PatternFill(start_color='FFEE08', end_color='FFEE08', fill_type = 'solid')
            ws['C'+str(cont)].fill = PatternFill(start_color='FFEE08', end_color='FFEE08', fill_type = 'solid')
            ws['D'+str(cont)].fill = PatternFill(start_color='FFEE08', end_color='FFEE08', fill_type = 'solid')
            ws['E'+str(cont)].fill = PatternFill(start_color='FFEE08', end_color='FFEE08', fill_type = 'solid')
            ws['F'+str(cont)].fill = PatternFill(start_color='FFEE08', end_color='FFEE08', fill_type = 'solid')
            ws['G'+str(cont)].fill = PatternFill(start_color='FFEE08', end_color='FFEE08', fill_type = 'solid')
            ws['H'+str(cont)].fill = PatternFill(start_color='FFEE08', end_color='FFEE08', fill_type = 'solid')
            cont += 1
            for detalle in DetallePedido.objects.filter(dtl_id_pedido=pedido['dtl_id_pedido']):
                ws.cell(row=cont, column=1).value = str(detalle.dtl_id_pedido)
                ws.cell(row=cont, column=2).value = str(detalle.dtl_id_pedido.ped_id_Suc)
                ws.cell(row=cont, column=3).value = str(detalle.dtl_id_pedido.ped_id_Suc.suc_zona)
                ws.cell(row=cont, column=4).value = str(detalle.dtl_codigo)
                ws.cell(row=cont, column=5).value = str(detalle.dtl_descripcion)
                ws.cell(row=cont, column=6).value = detalle.dtl_cantidad
                ws.cell(row=cont, column=7).value = detalle.dtl_precio
                ws.cell(row=cont, column=8).value = (detalle.dtl_cantidad * detalle.dtl_precio)
                cont += 1
        
        iva = sub_total_gobal*0.16
        total = sub_total_gobal + iva
        ws['G{}'.format(cont+1)] = 'Subtotal'
        ws['H{}'.format(cont+1)] = sub_total_gobal
        ws.cell(row=cont+1, column=8).number_format = '#,##0.00'

        ws['G{}'.format(cont+2)] = 'IVA'
        ws['H{}'.format(cont+2)] = iva
        ws.cell(row=cont+2, column=8).number_format = '#,##0.00'

        ws['G{}'.format(cont+3)] = 'TOTAL'
        ws['H{}'.format(cont+3)] = total
        ws.cell(row=cont+3, column=8).number_format = '#,##0.00'


        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    if cell.row != 1:
                        dims[cell.column] = max(
                            (dims.get(cell.column, 0), len(str(cell.value))))

        for col, value in dims.items():
            ws.column_dimensions[get_column_letter(col)].width = value+1

        
        nombre_archivo = 'Reporte_de_Pedidos_'+naturalday(inicio)+'_al_'+naturalday(fin)+'.xls'
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response

class pdf_reporte_gen(View):
    def get(self, request, *args, **kwargs):
        from django.contrib.humanize.templatetags.humanize import intcomma
        from django.contrib.humanize.templatetags.humanize import naturalday
        print ("Genero el PDF")
        response = HttpResponse(content_type='application/pdf')
        pdf_name = "clientes.pdf"  # llamado clientes
        # la linea 26 es por si deseas descargar el pdf a tu computadora
        inicio=self.request.GET.get('inicio')
        fin=self.request.GET.get('fin')
        tipo_pedido=self.request.GET.get('tipo_pedido')
        status=self.request.GET.getlist('status')
        query_resultado=Pedido.objects.filter(ped_fechaCreacion__range=(inicio, fin), ped_estatusPedido__in=status)
        if tipo_pedido != '0':
            query_resultado=query_resultado.filter(dtl_tipo_pedido=tipo_pedido)






        buff = BytesIO()
        doc = SimpleDocTemplate(buff,
                                pagesize=landscape(letter),
                                rightMargin=40,
                                leftMargin=40,
                                topMargin=60,
                                bottomMargin=18,
                                )
        items = []
        styles = getSampleStyleSheet()
        
        
        header = Paragraph("DETALLE DE PEDIDOS", styles['Heading1'])
        Paragraph("DETALLE DE PEDIDOS", styles['BodyText'])
        items.append(header)
    
        headings = ('N° Pedido', 'Fecha Pedido', 'Sucursal', 'Zona', 'Tipo', 'Estatus', 'Total')
        query_result = [(p.ped_id_ped, Paragraph(naturalday(p.ped_fechaCreacion), styles['BodyText']), Paragraph(str(p.ped_id_Suc), styles['BodyText']),Paragraph(str(p.ped_id_Suc.suc_zona), styles['BodyText']), Paragraph(p.get_dtl_tipo_pedido_display(), styles['BodyText']), Paragraph(p.get_ped_estatusPedido_display(), styles['BodyText']),intcomma(p.total_venta())) for p in query_resultado]
        
        suma_subtotal=0
        for item in query_resultado:
            suma_subtotal += item.total_venta()

        suma_subtotal = round(suma_subtotal, 2)
        IVA = round(suma_subtotal*0.16, 2)
        TOTAL =  round(suma_subtotal + IVA, 2)
        row_sub_total =[('','','','','','Subtotal',Paragraph(intcomma(suma_subtotal), styles['BodyText']))]
        row_sub_total+=[('','','','','','IVA',Paragraph(intcomma(IVA), styles['BodyText']))]
        row_sub_total +=[('','','','','','TOTAL',Paragraph(intcomma(TOTAL), styles['BodyText']))]
    
            

        

        t = Table([headings] + query_result+row_sub_total, colWidths=[2 * cm, 4 * cm, 5 * cm, 6 * cm, 4 * cm, 4 * cm, 2 * cm])
        t.setStyle(TableStyle(
            [
                ('GRID', (0, 0), (6, -1), 1, colors.dodgerblue),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
                ('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue)
            ]
        ))
        items.append(t)
        doc.build(items)
        response.write(buff.getvalue())
        buff.close()
        return response

